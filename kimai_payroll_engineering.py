#!/usr/bin/env python3
"""
Kimai Time Tracking Data Extraction for GitHub Actions
Extracts YTD timesheet data and creates Power BI ready Excel files
"""

import requests
import pandas as pd
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import time
import logging
import json
import sys
from typing import Dict, List, Optional

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Expected customers for Power BI consistency
EXPECTED_CUSTOMERS = [
    "Bluewater Operations",
    "Construction", 
    "DUKE",
    "Fire Dragon",
    "Flock",
    "General Admin",
    "Lumen",
    "Redspeed",
    "SCC- Miami Dade",
    "TECO",
    "TDS"
]

class KimaiExtractor:
    def __init__(self, api_token: str, base_url: str):
        self.api_config = {
            "headers": {
                "X-AUTH-USER": "API",
                "X-AUTH-TOKEN": api_token,
                "X-AUTH-PERMISSION": "view_user, view_other_timesheet, full"
            },
            "base_url": base_url.rstrip('/')
        }
        
    def calculate_ytd_dates(self):
        """Calculate YTD date range"""
        today = datetime.today()
        start_date = datetime(today.year, 1, 1, 0, 0, 0, 0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        logger.info(f"YTD Range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        return start_date, end_date

    def calculate_week_number(self, date):
        """Calculate week number (Sunday-based weeks)"""
        jan_1 = datetime(date.year, 1, 1).date()
        days_to_first_sunday = (6 - jan_1.weekday()) % 7
        if jan_1.weekday() == 6:  # If Jan 1 is Sunday
            days_to_first_sunday = 0
        first_sunday = jan_1 + timedelta(days=days_to_first_sunday)
        
        if date < first_sunday:
            return 1
        else:
            days_since_first_sunday = (date - first_sunday).days
            return (days_since_first_sunday // 7) + 1

    def fetch_api_data(self):
        """Fetch data from Kimai API"""
        start_date, end_date = self.calculate_ytd_dates()
        
        endpoints = [
            {
                "url": f"{self.api_config['base_url']}/api/timesheets",
                "params": {
                    "user": "all", 
                    "begin": start_date.strftime("%Y-%m-%dT%H:%M:%S"), 
                    "end": end_date.strftime("%Y-%m-%dT%H:%M:%S")
                },
                "name": "Timesheets",
                "paginated": True
            },
            {"url": f"{self.api_config['base_url']}/api/activities", "params": {"user": "all"}, "name": "Activities", "paginated": False},
            {"url": f"{self.api_config['base_url']}/api/projects", "params": {"user": "all"}, "name": "Projects", "paginated": False},
            {"url": f"{self.api_config['base_url']}/api/users", "params": {"user": "all", "visible": 3}, "name": "Users", "paginated": False}
        ]
        
        dataframes = {}
        
        for endpoint in endpoints:
            df_list = []
            
            if endpoint.get("paginated"):
                page = 1
                while True:
                    logger.info(f"Fetching {endpoint['name']} page {page}")
                    try:
                        response = requests.get(
                            endpoint["url"], 
                            headers=self.api_config["headers"], 
                            params={**endpoint["params"], "page": page},
                            timeout=30
                        )
                        
                        if response.status_code == 200:
                            data = response.json()
                            if isinstance(data, list) and data:
                                df_list.append(pd.DataFrame(data))
                                page += 1
                            else:
                                logger.info(f"No more data for {endpoint['name']} on page {page}")
                                break
                        else:
                            logger.error(f"Error fetching {endpoint['name']}: {response.status_code}")
                            break
                    except Exception as e:
                        logger.error(f"Exception fetching {endpoint['name']}: {e}")
                        break
            else:
                logger.info(f"Fetching {endpoint['name']}")
                try:
                    response = requests.get(
                        endpoint["url"], 
                        headers=self.api_config["headers"], 
                        params=endpoint["params"],
                        timeout=30
                    )
                    
                    if response.status_code == 200:
                        data = response.json()
                        if isinstance(data, list) and data:
                            df_list.append(pd.DataFrame(data))
                    else:
                        logger.error(f"Error fetching {endpoint['name']}: {response.status_code}")
                except Exception as e:
                    logger.error(f"Exception fetching {endpoint['name']}: {e}")
            
            if df_list:
                dataframes[endpoint["name"]] = pd.concat(df_list, ignore_index=True)
                logger.info(f"Loaded {len(dataframes[endpoint['name']])} records for {endpoint['name']}")
            else:
                logger.warning(f"No data retrieved for {endpoint['name']}")
        
        return dataframes

    def calculate_breaks_from_raw_data(self, raw_timesheets_df, users_df):
        """Calculate break durations from raw API data"""
        logger.info("Calculating break durations...")
        
        if raw_timesheets_df.empty or 'begin' not in raw_timesheets_df.columns:
            logger.warning("Missing timesheet data or time columns")
            return pd.DataFrame()
        
        # Create user mapping
        users_df["id"] = users_df["id"].astype(str)
        raw_timesheets_df["user"] = raw_timesheets_df["user"].astype(str)
        id_to_alias = dict(zip(users_df["id"], users_df["alias"]))
        
        # Convert timestamps
        raw_timesheets_df['begin_dt'] = pd.to_datetime(raw_timesheets_df['begin'], errors='coerce', utc=True)
        raw_timesheets_df['end_dt'] = pd.to_datetime(raw_timesheets_df['end'], errors='coerce', utc=True)
        raw_timesheets_df['begin_local'] = raw_timesheets_df['begin_dt'].dt.tz_convert('US/Eastern')
        raw_timesheets_df['end_local'] = raw_timesheets_df['end_dt'].dt.tz_convert('US/Eastern')
        raw_timesheets_df['date_only'] = raw_timesheets_df['begin_local'].dt.date
        
        raw_timesheets_df = raw_timesheets_df.dropna(subset=['begin_local', 'end_local'])
        
        break_records = []
        
        # Calculate breaks for each user/date
        for (user_id, date), group in raw_timesheets_df.groupby(['user', 'date_only']):
            user_alias = id_to_alias.get(user_id, f"User_{user_id}")
            group = group.sort_values('begin_local')
            
            if len(group) < 2:
                continue
            
            daily_breaks = []
            total_break_time = timedelta(0)
            
            # Find gaps between consecutive entries
            for i in range(len(group) - 1):
                current_end = group.iloc[i]['end_local']
                next_begin = group.iloc[i + 1]['begin_local']
                gap = next_begin - current_end
                
                # Valid breaks: 5 minutes to 1 hour, between 10 AM and 4 PM
                if 300 <= gap.total_seconds() <= 3600:
                    if 10 <= current_end.hour <= 16 and 10 <= next_begin.hour <= 16:
                        break_duration_hours = gap.total_seconds() / 3600
                        daily_breaks.append({
                            'break_start': current_end.strftime('%H:%M:%S'),
                            'break_end': next_begin.strftime('%H:%M:%S'),
                            'break_duration_hours': round(break_duration_hours, 2),
                            'break_duration_minutes': round(gap.total_seconds() / 60, 1)
                        })
                        total_break_time += gap
            
            if daily_breaks:
                break_records.append({
                    'user': user_alias,
                    'user_id': user_id,
                    'date': date,
                    'date_formatted': date.strftime('%m/%d/%Y'),
                    'total_breaks': len(daily_breaks),
                    'total_break_hours': round(total_break_time.total_seconds() / 3600, 2),
                    'total_break_minutes': round(total_break_time.total_seconds() / 60, 1),
                    'break_details': daily_breaks,
                    'first_work_start': group.iloc[0]['begin_local'].strftime('%H:%M:%S'),
                    'last_work_end': group.iloc[-1]['end_local'].strftime('%H:%M:%S'),
                    'total_work_entries': len(group)
                })
        
        logger.info(f"Found {len(break_records)} user-day combinations with breaks")
        return pd.DataFrame(break_records)

    def process_timesheet_mappings(self, dataframes):
        """Process timesheet data with ID mappings"""
        try:
            if not all(sheet in dataframes for sheet in ["Timesheets", "Activities", "Projects", "Users"]):
                logger.error("Missing required data sheets")
                return None, pd.DataFrame()
            
            timesheets_raw_df = dataframes["Timesheets"]
            activities_df = dataframes["Activities"] 
            projects_df = dataframes["Projects"]
            users_df = dataframes["Users"]
            
            # Calculate breaks
            breaks_df = self.calculate_breaks_from_raw_data(timesheets_raw_df.copy(), users_df.copy())
            
            # Process timesheets
            timesheets_df = timesheets_raw_df.copy()
            
            # Ensure consistent data types
            timesheets_df["project"] = timesheets_df["project"].astype(str)
            projects_df["id"] = projects_df["id"].astype(str)
            timesheets_df["user"] = timesheets_df["user"].astype(str)
            users_df["id"] = users_df["id"].astype(str)
            
            # Create mappings
            id_to_name = dict(zip(activities_df["id"], activities_df["name"]))
            id_to_project_name = dict(zip(projects_df["id"], projects_df["name"]))
            id_to_parent_title = dict(zip(projects_df["id"], projects_df["parentTitle"]))
            id_to_alias = dict(zip(users_df["id"], users_df["alias"]))
            
            # Apply mappings
            timesheets_df["activity"] = timesheets_df["activity"].map(id_to_name)
            timesheets_df["tags"] = timesheets_df["project"].map(id_to_parent_title)
            timesheets_df["project"] = timesheets_df["project"].map(id_to_project_name)
            timesheets_df["user"] = timesheets_df["user"].map(id_to_alias)
            
            # Format dates and calculate week numbers
            timesheets_df["begin"] = pd.to_datetime(timesheets_df["begin"], errors="coerce", utc=True)
            timesheets_df["begin_local"] = timesheets_df["begin"].dt.tz_convert('US/Eastern')
            
            # Calculate week number for each entry
            timesheets_df["Week"] = timesheets_df["begin_local"].apply(
                lambda x: self.calculate_week_number(x.date()) if pd.notna(x) else 1
            )
            
            # Format date for display
            timesheets_df["begin"] = timesheets_df["begin_local"].dt.strftime("%m/%d/%Y")
            
            # Convert duration to hours
            timesheets_df["duration"] = (timesheets_df["duration"] / 3600).round(2)
            
            # Add placeholder columns
            timesheets_df["price"] = 0.0
            timesheets_df["total"] = 0.0
            
            # Clean up columns
            if "rate" in timesheets_df.columns:
                timesheets_df = timesheets_df.drop(columns=["rate"])
            
            timesheets_df.rename(columns={"tags": "customer", "begin": "date"}, inplace=True)
            
            columns_to_remove = ["exported", "billable", "metaFields", "id", "end", "begin_dt", "begin_local"]
            timesheets_df = timesheets_df.drop(columns=[col for col in columns_to_remove if col in timesheets_df.columns])
            
            return timesheets_df, breaks_df
            
        except Exception as e:
            logger.error(f"Error processing timesheet mappings: {e}")
            return None, pd.DataFrame()

    def create_customer_reports(self, timesheets_df, breaks_df):
        """Create customer breakdown reports"""
        try:
            logger.info("Creating customer reports...")
            
            # Filter data
            df_customers = timesheets_df[timesheets_df["customer"] != "General Admin"]
            columns_to_keep = ["activity", "project", "user", "customer", "date", "duration"]
            if "Week" in df_customers.columns:
                columns_to_keep.append("Week")
            df_customers = df_customers[columns_to_keep]
            
            # Convert data types
            df_customers["duration"] = pd.to_numeric(df_customers["duration"], errors="coerce").fillna(0)
            df_customers["date"] = pd.to_datetime(df_customers["date"], errors="coerce")
            
            # Use all YTD data
            df_ytd = df_customers
            
            # Group data
            grouped_total_ytd = df_ytd.groupby(["customer", "user"]).agg({"duration": "sum"}).reset_index()
            grouped_detailed_ytd = df_ytd.groupby(["customer", "user", "project", "activity", "date"]).agg({"duration": "sum"}).reset_index()
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if wb.active:
                wb.remove(wb.active)
            
            # Create Employee Summary
            ws_emp_summary = wb.create_sheet(title="Employee Summary", index=0)
            ws_emp_summary.append(["Customer", "Employee", "TotalHours", "Date"])
            
            for customer in grouped_total_ytd["customer"].unique():
                customer_total = grouped_total_ytd[grouped_total_ytd["customer"] == customer]
                for _, row in customer_total.iterrows():
                    ws_emp_summary.append([
                        customer,
                        row["user"],
                        round(float(row["duration"]), 2),
                        datetime.now().strftime('%m/%d/%Y')
                    ])
            
            # Create Activity Details
            ws_activities = wb.create_sheet(title="Activity Details", index=1)
            ws_activities.append(["Customer", "Employee", "Project", "Activity", "Hours", "Date"])
            
            for customer in grouped_detailed_ytd["customer"].unique():
                customer_detailed = grouped_detailed_ytd[grouped_detailed_ytd["customer"] == customer]
                for _, row in customer_detailed.iterrows():
                    actual_date = row["date"].strftime('%m/%d/%Y') if pd.notna(row["date"]) else datetime.now().strftime('%m/%d/%Y')
                    ws_activities.append([
                        customer,
                        row["user"],
                        row["project"],
                        row["activity"],
                        round(float(row["duration"]), 2),
                        actual_date
                    ])
            
            # Create Customer Hours Summary
            ws_pie = wb.create_sheet(title="Customer Hours Summary", index=2)
            ws_pie.append(["Customer", "TotalHours", "Percentage"])
            
            customer_hours = df_ytd.groupby("customer")["duration"].sum().reset_index()
            customer_hours = customer_hours.sort_values("duration", ascending=False)
            total_all_customers = customer_hours["duration"].sum()
            
            for _, row in customer_hours.iterrows():
                ws_pie.append([
                    row["customer"], 
                    round(float(row["duration"]), 2), 
                    round((row["duration"] / total_all_customers) * 100, 1) if total_all_customers > 0 else 0
                ])
            
            # Create customer sheets
            customers_with_data = set()
            if not grouped_total_ytd.empty:
                customers_with_data = set(grouped_total_ytd["customer"].dropna().unique())
            
            all_expected_customers = set(EXPECTED_CUSTOMERS) - {"General Admin"}
            
            for customer in all_expected_customers:
                if pd.isna(customer) or customer is None:
                    continue
                    
                customer_name = str(customer)[:31]
                
                if customer in customers_with_data:
                    # Create sheet with data
                    customer_total = grouped_total_ytd[grouped_total_ytd["customer"] == customer]
                    customer_detailed = df_ytd[df_ytd["customer"] == customer].groupby(["user", "project", "activity"]).agg({"duration": "sum"}).reset_index()
                    
                    ws = wb.create_sheet(title=customer_name)
                    ws.append(["EmployeeName", "EmployeeHours", "ProjectName", "ActivityName", "ActivityHours", "CustomerTotal"])
                    
                    customer_total_hours = round(float(customer_total["duration"].sum()), 2)
                    
                    for _, emp_row in customer_total.iterrows():
                        employee = emp_row["user"]
                        total_hours = round(float(emp_row["duration"]), 2)
                        
                        emp_activities = customer_detailed[customer_detailed["user"] == employee]
                        
                        if len(emp_activities) > 0:
                            for _, act_row in emp_activities.iterrows():
                                ws.append([
                                    str(employee),
                                    float(total_hours),
                                    str(act_row["project"]),
                                    str(act_row["activity"]),
                                    float(round(act_row["duration"], 2)),
                                    float(customer_total_hours)
                                ])
                        else:
                            ws.append([str(employee), float(total_hours), "", "", 0.0, float(customer_total_hours)])
                else:
                    # Create empty sheet
                    ws = wb.create_sheet(title=customer_name)
                    ws.append(["EmployeeName", "EmployeeHours", "ProjectName", "ActivityName", "ActivityHours", "CustomerTotal"])
                    ws.append(["", 0.0, "", "", 0.0, 0.0])
            
            # Create break analysis sheet
            if not breaks_df.empty:
                ws_breaks = wb.create_sheet(title="YTD Break Analysis", index=3)
                ws_breaks.append(["Employee", "Date", "BreakStart", "BreakEnd", "DurationMinutes", "DurationHours", "WorkDayStart", "WorkDayEnd"])
                
                # Process breaks data for individual entries
                individual_breaks = []
                for _, row in breaks_df.iterrows():
                    if 'break_details' in row and isinstance(row['break_details'], list):
                        for break_info in row['break_details']:
                            individual_breaks.append({
                                'user': row['user'],
                                'date': row['date_formatted'],
                                'break_start_time': break_info['break_start'],
                                'break_end_time': break_info['break_end'],
                                'break_duration_minutes': break_info['break_duration_minutes'],
                                'break_duration_hours': break_info['break_duration_hours'],
                                'work_day_start': row['first_work_start'],
                                'work_day_end': row['last_work_end']
                            })
                
                individual_breaks_df = pd.DataFrame(individual_breaks)
                if not individual_breaks_df.empty:
                    individual_breaks_sorted = individual_breaks_df.sort_values(['user', 'date', 'break_start_time'])
                    for _, row in individual_breaks_sorted.iterrows():
                        ws_breaks.append([
                            str(row['user']),
                            str(row['date']),
                            str(row['break_start_time']),
                            str(row['break_end_time']),
                            float(round(row['break_duration_minutes'], 1)),
                            float(round(row['break_duration_hours'], 2)),
                            str(row['work_day_start']),
                            str(row['work_day_end'])
                        ])
            
            # Save main file
            main_filename = f"Kimai_Customer_Reports_YTD_{datetime.now().strftime('%Y%m%d')}.xlsx"
            wb.save(main_filename)
            logger.info(f"Customer reports saved to {main_filename}")
            
            # Create summary file
            self.create_summary_file(timesheets_df)
            
            return main_filename
            
        except Exception as e:
            logger.error(f"Error creating customer reports: {e}")
            return None

    def create_summary_file(self, timesheets_df):
        """Create summary file"""
        try:
            # Group by user for summary
            df_hours_per_employee = timesheets_df.groupby('user').agg({'duration': 'sum'}).reset_index()
            
            # Create summary workbook
            wb_summary = Workbook()
            if wb_summary.active:
                wb_summary.remove(wb_summary.active)
            
            # Create Weekly Summary sheet
            ws_summary = wb_summary.create_sheet(title="Weekly Summary", index=0)
            
            # Add employee data (no headers for consistency)
            for _, row in df_hours_per_employee.iterrows():
                ws_summary.append([row['user'], round(float(row['duration']), 2)])
            
            # Add summary metrics
            total_hours = df_hours_per_employee['duration'].sum()
            total_employees = len(df_hours_per_employee)
            
            ws_summary.append([])
            ws_summary.append(["METRIC", "VALUE"])
            ws_summary.append(["Total Hours", round(total_hours, 2)])
            ws_summary.append(["Total Employees", total_employees])
            ws_summary.append(["Average Hours", round(total_hours / total_employees, 2) if total_employees > 0 else 0])
            
            summary_filename = f"Kimai_Summary_YTD_{datetime.now().strftime('%Y%m%d')}.xlsx"
            wb_summary.save(summary_filename)
            logger.info(f"Summary saved to {summary_filename}")
            
        except Exception as e:
            logger.error(f"Error creating summary: {e}")

def main():
    """Main execution function"""
    try:
        # Get credentials from environment
        api_token = os.getenv('KIMAI_API_TOKEN')
        base_url = os.getenv('KIMAI_BASE_URL', 'https://time.mybwt.net')
        
        # Debug authentication
        print(f"=== DEBUG AUTHENTICATION ===")
        print(f"API Token: {api_token}")
        print(f"Base URL: {base_url}")
        print(f"API Token exists: {'KIMAI_API_TOKEN' in os.environ}")
        print(f"Base URL exists: {'KIMAI_BASE_URL' in os.environ}")
        
        if not api_token:
            logger.error("Missing KIMAI_API_TOKEN environment variable")
            sys.exit(1)
        
        # Test simple API call before proceeding
        test_headers = {
            "X-AUTH-USER": "API",
            "X-AUTH-TOKEN": api_token,
            "X-AUTH-PERMISSION": "view_user, view_other_timesheet, full",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "DNT": "1",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1"
        }
        
        test_url = f"{base_url}/api/users"
        print(f"Testing URL: {test_url}")
        print(f"Test headers: {test_headers}")
        
        try:
            response = requests.get(test_url, headers=test_headers, timeout=10)
            print(f"Response status: {response.status_code}")
            print(f"Response headers: {dict(response.headers)}")
            
            if response.status_code != 200:
                print(f"Response text: {response.text[:500]}")
                
                # Try alternative authentication methods
                print("=== TRYING ALTERNATIVE AUTH ===")
                
                # Method 1: Different header format
                alt_headers_1 = {
                    "Authorization": f"Bearer {api_token}",
                    "Content-Type": "application/json"
                }
                alt_response_1 = requests.get(test_url, headers=alt_headers_1, timeout=10)
                print(f"Alt Method 1 (Bearer): {alt_response_1.status_code}")
                
                # Method 2: Basic auth
                alt_headers_2 = {
                    "X-AUTH-TOKEN": api_token,
                    "Content-Type": "application/json"
                }
                alt_response_2 = requests.get(test_url, headers=alt_headers_2, timeout=10)
                print(f"Alt Method 2 (Token only): {alt_response_2.status_code}")
                
            else:
                print("Authentication test successful!")
                
        except Exception as e:
            print(f"Test request failed: {str(e)}")
        
        print("=== END DEBUG ===")
        
        logger.info(f"Starting Kimai YTD extraction from {base_url}")
        
        # Create extractor
        extractor = KimaiExtractor(api_token, base_url)
        
        # Fetch data
        logger.info("Fetching data from Kimai API...")
        dataframes = extractor.fetch_api_data()
        
        if not dataframes:
            logger.error("No data retrieved from Kimai API")
            return False
        
        # Process data
        logger.info("Processing timesheet data...")
        timesheets_df, breaks_df = extractor.process_timesheet_mappings(dataframes)
        
        if timesheets_df is None:
            logger.error("Failed to process timesheet data")
            return False
        
        # Create reports
        logger.info("Creating customer reports...")
        main_file = extractor.create_customer_reports(timesheets_df, breaks_df)
        
        if main_file:
            total_records = len(timesheets_df)
            logger.info(f"Kimai extraction completed: {total_records} timesheet records processed")
            logger.info(f"Files created: {main_file} and summary file")
            return True
        else:
            logger.error("Failed to create reports")
            return False
        
    except Exception as e:
        logger.error(f"Kimai extraction failed: {str(e)}")
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
